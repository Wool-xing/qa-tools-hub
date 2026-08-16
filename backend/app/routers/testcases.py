"""Test Case Manager — CRUD API for QA通关."""

import io
from datetime import datetime, timezone
from fastapi import APIRouter, Depends, HTTPException, Query, Response, UploadFile, File
from fastapi.responses import StreamingResponse
from sqlalchemy import select, func, or_
from sqlalchemy.ext.asyncio import AsyncSession
from pydantic import BaseModel, field_validator
import openpyxl
from app.database import get_db
from app.models.user import User
from app.models.test_case import TestCase
from app.models.test_run import TestRun
from app.routers.auth import get_current_user
from app.config import MAX_XLSX_BYTES, BULK_UPDATE_LIMIT

router = APIRouter(prefix="/api/testcases", tags=["testcases"])

# Limit moved to config — imported above
_MAX_XLSX_BYTES = MAX_XLSX_BYTES


def _escape_csv_cell(value: str) -> str:
    """Sanitize cell value: prevent CSV formula injection + strip HTML tags."""
    import re as _re
    if not value:
        return value
    # Strip HTML tags and decode entities
    cleaned = _re.sub(r'<[^>]*>', '', value)
    cleaned = cleaned.replace('&lt;', '<').replace('&gt;', '>').replace('&amp;', '&')
    # Prefix formula-trigger chars
    if cleaned and cleaned[0] in "=+-@":
        return "'" + cleaned
    return cleaned


class TestCaseCreate(BaseModel):
    title: str
    steps: str = ""
    expected_result: str = ""
    priority: str = "P2"
    status: str = "draft"
    tags: str = ""
    folder: str = "默认"
    level_id: int | None = None
    team_id: int | None = None

    @field_validator("title")
    @classmethod
    def check_title(cls, v: str) -> str:
        if not v or not v.strip():
            raise ValueError("Title must not be empty")
        return v.strip()

    @field_validator("folder")
    @classmethod
    def check_folder(cls, v: str) -> str:
        return v.strip() if v.strip() else "默认"

    @field_validator("priority")
    @classmethod
    def check_priority(cls, v: str) -> str:
        if v.upper() not in {"P0", "P1", "P2", "P3", "P4"}:
            raise ValueError("Priority must be P0-P4")
        return v.upper()

    @field_validator("status")
    @classmethod
    def check_status(cls, v: str) -> str:
        if v.lower() not in {"draft", "ready", "running", "passed", "failed"}:
            raise ValueError("Status must be draft/ready/running/passed/failed")
        return v.lower()


class TestCaseUpdate(BaseModel):
    title: str | None = None
    steps: str | None = None
    expected_result: str | None = None
    priority: str | None = None
    status: str | None = None
    tags: str | None = None
    folder: str | None = None
    level_id: int | None = None
    team_id: int | None = None


@router.get("")
async def list_testcases(
    response: Response,
    search: str = Query(""),
    folder: str = Query(""),
    priority: str = Query(""),
    status: str = Query(""),
    limit: int = Query(50, ge=1, le=200),
    offset: int = Query(0, ge=0),
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    # Count total first
    count_q = select(func.count(TestCase.id)).where(TestCase.user_id == user.id)
    if search:
        count_q = count_q.where(or_(TestCase.title.contains(search), TestCase.steps.contains(search), TestCase.tags.contains(search)))
    if folder:
        count_q = count_q.where(TestCase.folder == folder)
    if priority:
        count_q = count_q.where(TestCase.priority == priority)
    if status:
        count_q = count_q.where(TestCase.status == status)
    total = (await db.execute(count_q)).scalar()

    q = select(TestCase).where(TestCase.user_id == user.id)
    if search:
        q = q.where(or_(TestCase.title.contains(search), TestCase.steps.contains(search), TestCase.tags.contains(search)))
    if folder:
        q = q.where(TestCase.folder == folder)
    if priority:
        q = q.where(TestCase.priority == priority)
    if status:
        q = q.where(TestCase.status == status)
    q = q.order_by(TestCase.updated_at.desc()).offset(offset).limit(limit)
    result = await db.execute(q)
    cases = result.scalars().all()

    # Get folder counts (apply search/priority/status filters, not folder filter)
    folder_q = select(TestCase.folder, func.count(TestCase.id)).where(TestCase.user_id == user.id)
    if search:
        folder_q = folder_q.where(or_(TestCase.title.contains(search), TestCase.steps.contains(search), TestCase.tags.contains(search)))
    if priority:
        folder_q = folder_q.where(TestCase.priority == priority)
    if status:
        folder_q = folder_q.where(TestCase.status == status)
    folder_q = folder_q.group_by(TestCase.folder)
    folder_result = await db.execute(folder_q)
    folders = {row[0]: row[1] for row in folder_result.all()}

    response.headers["X-Total-Count"] = str(total)
    return {
        "cases": [{
            "id": c.id, "title": c.title, "steps": c.steps, "expected_result": c.expected_result,
            "priority": c.priority, "status": c.status, "tags": c.tags or "",
            "folder": c.folder, "created_at": c.created_at.isoformat() if c.created_at else None,
            "updated_at": c.updated_at.isoformat() if c.updated_at else None,
        } for c in cases],
        "folders": [{"name": k, "count": v} for k, v in sorted(folders.items())],
        "total": total,
    }


@router.post("")
async def create_testcase(data: TestCaseCreate, user: User = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    tc = TestCase(user_id=user.id, title=data.title, steps=data.steps, expected_result=data.expected_result,
                  priority=data.priority, status=data.status, tags=data.tags, folder=data.folder,
                  level_id=data.level_id, team_id=data.team_id)
    db.add(tc); await db.commit(); await db.refresh(tc)
    return {"id": tc.id, "title": tc.title, "folder": tc.folder}


@router.put("/{tc_id}")
async def update_testcase(tc_id: int, data: TestCaseUpdate, user: User = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    r = await db.execute(select(TestCase).where(TestCase.id == tc_id, TestCase.user_id == user.id))
    tc = r.scalar_one_or_none()
    if not tc:
        raise HTTPException(status_code=404)
    for k, v in data.model_dump(exclude_unset=True).items():
        setattr(tc, k, v)
    tc.updated_at = datetime.now(timezone.utc)
    await db.commit()
    return {"id": tc.id, "updated": True}


@router.delete("/{tc_id}")
async def delete_testcase(tc_id: int, user: User = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    r = await db.execute(select(TestCase).where(TestCase.id == tc_id, TestCase.user_id == user.id))
    tc = r.scalar_one_or_none()
    if not tc:
        raise HTTPException(status_code=404)
    await db.delete(tc); await db.commit()
    return {"id": tc_id, "deleted": True}


# ==================== Test Runs ====================


class TestRunCreate(BaseModel):
    status: str = "passed"
    notes: str = ""


@router.post("/{tc_id}/runs")
async def add_test_run(tc_id: int, data: TestRunCreate, user: User = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    r = await db.execute(select(TestCase).where(TestCase.id == tc_id, TestCase.user_id == user.id))
    if not r.scalar_one_or_none():
        raise HTTPException(status_code=404)
    tr = TestRun(test_case_id=tc_id, user_id=user.id, status=data.status, notes=data.notes)
    db.add(tr); await db.commit(); await db.refresh(tr)
    # Update test case status to match
    tc = (await db.execute(select(TestCase).where(TestCase.id == tc_id))).scalar_one()
    tc.status = data.status; tc.updated_at = datetime.now(timezone.utc); await db.commit()
    return {"id": tr.id, "status": tr.status, "created_at": tr.created_at.isoformat() if tr.created_at else None}


@router.get("/{tc_id}/runs")
async def list_test_runs(tc_id: int, user: User = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    r = await db.execute(select(TestRun).where(TestRun.test_case_id == tc_id, TestRun.user_id == user.id).order_by(TestRun.created_at.desc()).limit(20))
    runs = r.scalars().all()
    return [{"id": run.id, "status": run.status, "notes": run.notes, "created_at": run.created_at.isoformat() if run.created_at else None} for run in runs]


# ==================== Bulk Operations ====================

class BulkUpdate(BaseModel):
    ids: list[int]
    status: str | None = None
    folder: str | None = None

    @field_validator("ids")
    @classmethod
    def check_ids(cls, v: list[int]) -> list[int]:
        if len(v) > BULK_UPDATE_LIMIT:
            raise ValueError(f"Bulk update limited to {BULK_UPDATE_LIMIT} items per request")
        return v


@router.post("/bulk")
async def bulk_update(data: BulkUpdate, user: User = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    if not data.ids:
        return {"updated": 0}
    r = await db.execute(select(TestCase).where(TestCase.id.in_(data.ids), TestCase.user_id == user.id))
    cases = r.scalars().all()
    now = datetime.now(timezone.utc)
    updated = 0
    for tc in cases:
        if data.status:
            tc.status = data.status
        if data.folder:
            tc.folder = data.folder
        tc.updated_at = now
        updated += 1
    await db.commit()
    return {"updated": updated}


# ==================== Export ====================

@router.get("/export/csv")
async def export_csv(user: User = Depends(get_current_user), db: AsyncSession = Depends(get_db)):
    r = await db.execute(select(TestCase).where(TestCase.user_id == user.id).order_by(TestCase.folder, TestCase.id))
    cases = r.scalars().all()

    def generate():
        yield "ID,Title,Priority,Status,Folder,Tags,Steps,Expected\n"
        for c in cases:
            title = _escape_csv_cell(c.title.replace('"', '""'))
            steps = _escape_csv_cell((c.steps or "").replace('"', '""').replace('\n', '; '))
            exp = _escape_csv_cell((c.expected_result or "").replace('"', '""'))
            tags = _escape_csv_cell((c.tags or "").replace('"', '""'))
            folder = _escape_csv_cell(c.folder or "")
            priority = _escape_csv_cell(c.priority or "")
            status = _escape_csv_cell(c.status or "")
            yield f'{c.id},"{title}","{priority}","{status}","{folder}","{tags}","{steps}","{exp}"\n'

    return StreamingResponse(
        generate(),
        media_type="text/csv",
        headers={"Content-Disposition": "attachment; filename=testcases.csv"}
    )


# xlsx export
@router.get("/export/xlsx")
async def export_testcases_xlsx(
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    result = await db.execute(
        select(TestCase).where(TestCase.user_id == user.id).order_by(TestCase.updated_at.desc())
    )
    cases = result.scalars().all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Test Cases"

    # Header style
    header_font = openpyxl.styles.Font(bold=True, color="FFFFFF")
    header_fill = openpyxl.styles.PatternFill(start_color="6366F1", end_color="6366F1", fill_type="solid")
    header_alignment = openpyxl.styles.Alignment(horizontal="center", vertical="center")

    headers = ["ID", "Title", "Steps", "Expected Result", "Priority", "Status", "Tags", "Folder", "Created At", "Updated At"]
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment

    for row, c in enumerate(cases, 2):
        ws.cell(row=row, column=1, value=c.id)
        ws.cell(row=row, column=2, value=_escape_csv_cell(c.title))
        ws.cell(row=row, column=3, value=_escape_csv_cell(c.steps))
        ws.cell(row=row, column=4, value=_escape_csv_cell(c.expected_result))
        ws.cell(row=row, column=5, value=c.priority)
        ws.cell(row=row, column=6, value=c.status)
        ws.cell(row=row, column=7, value=_escape_csv_cell(c.tags or ""))
        ws.cell(row=row, column=8, value=_escape_csv_cell(c.folder or ""))
        ws.cell(row=row, column=9, value=c.created_at.isoformat() if c.created_at else "")
        ws.cell(row=row, column=10, value=c.updated_at.isoformat() if c.updated_at else "")

    # Column widths
    widths = [6, 30, 40, 40, 10, 10, 20, 15, 20, 20]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = "A2"

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=testcases.xlsx"}
    )


# xlsx import
class XlsxImportResult(BaseModel):
    created: int = 0
    skipped: int = 0
    errors: list[str] = []


@router.post("/import/xlsx", response_model=XlsxImportResult)
async def import_testcases_xlsx(
    file: UploadFile = File(...),
    user: User = Depends(get_current_user),
    db: AsyncSession = Depends(get_db),
):
    result = XlsxImportResult()
    try:
        contents = await file.read()
        if len(contents) > _MAX_XLSX_BYTES:
            result.errors.append(f"File size {len(contents)} bytes exceeds {_MAX_XLSX_BYTES} byte limit")
            return result
        wb = openpyxl.load_workbook(io.BytesIO(contents))
        ws = wb.active
        rows = list(ws.iter_rows(min_row=2, values_only=True))  # skip header
    except Exception as e:
        result.errors.append(f"Failed to parse xlsx: {str(e)}")
        return result

    for row_idx, row in enumerate(rows, 2):
        if not row or not any(row):
            continue
        try:
            title = str(row[1]).strip() if len(row) > 1 and row[1] else ""
            if not title:
                result.skipped += 1
                continue

            steps = str(row[2]).strip() if len(row) > 2 and row[2] else ""
            expected = str(row[3]).strip() if len(row) > 3 and row[3] else ""
            priority = str(row[4]).strip().upper() if len(row) > 4 and row[4] else "P2"
            status = str(row[5]).strip().lower() if len(row) > 5 and row[5] else "draft"
            tags = str(row[6]).strip() if len(row) > 6 and row[6] else ""
            folder = str(row[7]).strip() if len(row) > 7 and row[7] else ""

            # Validate priority
            valid_priorities = {"P0", "P1", "P2", "P3", "P4"}
            if priority not in valid_priorities:
                priority = "P2"

            # Validate status
            valid_statuses = {"draft", "ready", "running", "passed", "failed"}
            if status not in valid_statuses:
                status = "draft"

            tc = TestCase(
                user_id=user.id, title=title, steps=steps, expected_result=expected,
                priority=priority, status=status, tags=tags, folder=folder
            )
            db.add(tc)
            result.created += 1
            if result.created % 100 == 0:
                await db.commit()
        except Exception as e:
            result.errors.append(f"Row {row_idx}: {str(e)}")

    await db.commit()
    return result
